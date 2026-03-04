"""
Excel parser using openpyxl.
헤더 행을 읽어 컬럼을 자동 매핑 (컬럼 순서 변경에 안전).

지원 헤더 → 필드 매핑:
  심의번호/처리번호        → case_number
  심의의견                 → opinion_note
  작성일/처리일자          → case_date
  위반내용/심의지적코드    → violation_type
  한정표현                 → limit_expression
"""

from pathlib import Path

from openpyxl import load_workbook


# 헤더 텍스트 → 내부 필드명 매핑 (유사 표현 포함)
_HEADER_MAP: dict[str, str] = {
    "심의번호": "case_number",
    "처리번호": "case_number",
    "심의의견": "opinion_note",
    "작성일": "case_date",
    "처리일자": "case_date",
    "위반내용": "violation_type",
    "심의지적코드": "violation_type",
    "한정표현": "limit_expression",
}


def _detect_column_mapping(header_row: tuple) -> dict[int, str]:
    """헤더 행에서 컬럼 인덱스 → 필드명 매핑을 자동 생성."""
    mapping: dict[int, str] = {}
    for col_idx, cell in enumerate(header_row):
        header_text = str(cell or "").strip()
        if header_text in _HEADER_MAP:
            mapping[col_idx] = _HEADER_MAP[header_text]
    return mapping


class ExcelParser:
    """Parse an Excel (.xlsx) file; returns per-row dict with separate columns."""

    @staticmethod
    def parse(file_path: str) -> list[dict]:
        """
        Returns:
            List of {
                "case_number": str,       # 심의번호/처리번호
                "case_date": str,         # 작성일/처리일자
                "opinion_note": str,      # 심의의견 원문
                "violation_type": str,    # 위반내용/심의지적코드
                "limit_expression": str,  # 한정표현
                "row": int,
                "sheet": str,
                "source_file": str,
            }
        """
        path = Path(file_path)
        wb = load_workbook(str(path), read_only=True, data_only=True)

        results: list[dict] = []
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)

                # 1행: 헤더 읽기 → 자동 매핑
                header_row = next(rows_iter, None)
                if header_row is None:
                    continue
                col_map = _detect_column_mapping(header_row)

                if not col_map:
                    # 헤더가 없는 시트 → 스킵
                    continue

                # 2행~: 데이터
                for row_idx, row in enumerate(rows_iter, start=2):
                    record: dict[str, str] = {
                        "case_number": "",
                        "opinion_note": "",
                        "case_date": "",
                        "violation_type": "",
                        "limit_expression": "",
                    }

                    for col_idx, field_name in col_map.items():
                        if col_idx < len(row):
                            record[field_name] = str(row[col_idx] or "").strip()

                    # 심의번호·심의의견 중 하나라도 있어야 유효
                    if not record["case_number"] and not record["opinion_note"]:
                        continue

                    # 처리일자: datetime 객체일 경우 날짜 부분만 추출
                    if " " in record["case_date"]:
                        record["case_date"] = record["case_date"].split(" ")[0]

                    # 처리번호: 소수점 제거 (openpyxl이 숫자로 읽는 경우)
                    cn = record["case_number"]
                    if "." in cn:
                        record["case_number"] = cn.split(".")[0]

                    record["row"] = row_idx
                    record["sheet"] = sheet_name
                    record["source_file"] = path.name

                    results.append(record)
        finally:
            wb.close()
        return results